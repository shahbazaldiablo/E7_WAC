#!/usr/bin/env python3
"""
Eagle7 Website HTTP Status Crawler
Run:
    pip install requests beautifulsoup4 openpyxl
    python website_status_crawler.py
or:
    python website_status_crawler.py https://example.com
"""

import csv
import re
import sys
import time
from collections import defaultdict, deque
from datetime import datetime
from urllib.parse import urljoin, urlparse, urldefrag
from urllib.robotparser import RobotFileParser

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DEFAULT_URL = "https://indyroofandrestoration.com/"
USER_AGENT = "Eagle7 Website Status Crawler/1.0"
TIMEOUT = 20
DELAY = 0.15
MAX_PAGES = 5000
MAX_RESOURCES = 20000


def clean_url(url):
    url, _ = urldefrag(url.strip())
    return url


def same_domain(url, netloc):
    p = urlparse(url)
    return p.scheme in ("http", "https") and p.netloc.lower() == netloc.lower()


def category(status):
    if status == 200:
        return "200 OK"
    if status == 0:
        return "Connection Errors"
    if 300 <= status < 400:
        return "3xx Redirects"
    if 400 <= status < 500:
        return "4xx Client Errors"
    if 500 <= status < 600:
        return "5xx Server Errors"
    return "Other"


def css_urls(text):
    return [
        x.strip().strip("'\"")
        for x in re.findall(r"url\(\s*([^)]*)\)", text or "", re.I)
        if x.strip() and not x.strip().startswith("data:")
    ]


def discover(page_url, html):
    soup = BeautifulSoup(html, "html.parser")
    found = []

    def add(raw, kind):
        if not raw:
            return
        raw = raw.strip()
        if raw.startswith(("data:", "javascript:", "mailto:", "tel:", "#")):
            return
        url = clean_url(urljoin(page_url, raw))
        if urlparse(url).scheme in ("http", "https"):
            found.append((url, kind))

    for tag in soup.find_all("a", href=True):
        add(tag["href"], "Link")

    for tag in soup.find_all("img"):
        add(tag.get("src"), "Image")
        for item in (tag.get("srcset") or "").split(","):
            if item.strip():
                add(item.strip().split()[0], "Image srcset")

    for tag in soup.find_all("source"):
        add(tag.get("src"), "Source")
        for item in (tag.get("srcset") or "").split(","):
            if item.strip():
                add(item.strip().split()[0], "Source srcset")

    for tag in soup.find_all("script", src=True):
        add(tag["src"], "Script")

    for tag in soup.find_all("link", href=True):
        rel = " ".join(tag.get("rel", [])).lower()
        add(tag["href"], "Stylesheet" if "stylesheet" in rel else "Link resource")

    for tag in soup.find_all(style=True):
        for raw in css_urls(tag.get("style")):
            add(raw, "CSS URL")

    for raw in css_urls(html):
        add(raw, "CSS URL")

    return found


def request(session, url):
    started = time.perf_counter()
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True, stream=True)
        result = {
            "status": r.status_code,
            "final_url": r.url,
            "content_type": r.headers.get("Content-Type", ""),
            "response_time": round(time.perf_counter() - started, 3),
            "error": "",
        }
        r.close()
        return result
    except requests.RequestException as e:
        return {
            "status": 0,
            "final_url": "",
            "content_type": "",
            "response_time": round(time.perf_counter() - started, 3),
            "error": str(e),
        }


def get_html(session, url):
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
        status = r.status_code
        final = r.url
        content_type = r.headers.get("Content-Type", "")
        html = r.text if "text/html" in content_type.lower() else None
        r.close()
        return html, final, status, content_type
    except requests.RequestException:
        return None, "", 0, ""


def robots_parser(session, root):
    rp = RobotFileParser()
    try:
        r = session.get(urljoin(root, "/robots.txt"), timeout=TIMEOUT)
        if r.status_code == 200:
            rp.parse(r.text.splitlines())
            return rp
    except requests.RequestException:
        pass
    return None


def write_sheet(ws, rows):
    headers = [
        "Page URL", "Resource URL", "Resource Type", "HTTP Status",
        "Status Category", "Final URL", "Content-Type",
        "Response Time (s)", "Error"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = [45, 65, 22, 13, 22, 65, 35, 20, 55]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width


def main():
    root = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_URL
    if not root.startswith(("http://", "https://")):
        root = "https://" + root
    root = clean_url(root).rstrip("/") + "/"

    netloc = urlparse(root).netloc
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_domain = re.sub(r"[^A-Za-z0-9.-]", "_", netloc)
    xlsx_file = f"{safe_domain}_website_status_audit_{stamp}.xlsx"
    csv_file = xlsx_file.replace(".xlsx", ".csv")

    session = requests.Session()
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "*/*",
        "Accept-Language": "en-GB,en;q=0.9",
    })

    robots = robots_parser(session, root)
    queue = deque([root])
    discovered_pages = {root}
    checked_pages = set()
    checked_resources = set()
    rows = []

    print(f"\nEagle7 Website Status Crawler")
    print(f"Target: {root}\n")

    while queue and len(checked_pages) < MAX_PAGES:
        page = queue.popleft()
        if page in checked_pages:
            continue

        if robots and not robots.can_fetch(USER_AGENT, page):
            print(f"[ROBOTS] {page}")
            checked_pages.add(page)
            continue

        print(f"[PAGE {len(checked_pages) + 1}] {page}")
        html, final_page, status, content_type = get_html(session, page)
        checked_pages.add(page)

        page_result = request(session, page)
        rows.append({
            "Page URL": page,
            "Resource URL": page,
            "Resource Type": "Page",
            "HTTP Status": page_result["status"],
            "Status Category": category(page_result["status"]),
            "Final URL": page_result["final_url"],
            "Content-Type": page_result["content_type"],
            "Response Time (s)": page_result["response_time"],
            "Error": page_result["error"],
        })

        if html is None:
            time.sleep(DELAY)
            continue

        for resource, kind in discover(final_page or page, html):
            if kind == "Link" and same_domain(resource, netloc):
                if resource not in discovered_pages:
                    discovered_pages.add(resource)
                    queue.append(resource)

            key = (resource, kind)
            if key in checked_resources:
                continue
            if len(checked_resources) >= MAX_RESOURCES:
                break

            checked_resources.add(key)
            result = request(session, resource)

            rows.append({
                "Page URL": page,
                "Resource URL": resource,
                "Resource Type": kind,
                "HTTP Status": result["status"],
                "Status Category": category(result["status"]),
                "Final URL": result["final_url"],
                "Content-Type": result["content_type"],
                "Response Time (s)": result["response_time"],
                "Error": result["error"],
            })
            time.sleep(DELAY)

        time.sleep(DELAY)

    grouped = defaultdict(list)
    for row in rows:
        grouped[row["Status Category"]].append(row)

    order = [
        "200 OK", "3xx Redirects", "4xx Client Errors",
        "5xx Server Errors", "Connection Errors", "Other"
    ]

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["Website", root])
    summary.append(["Scan timestamp", stamp])
    summary.append(["Pages crawled", len(checked_pages)])
    summary.append(["Unique resources checked", len(checked_resources)])
    summary.append(["Total HTTP checks", len(rows)])
    summary.append([])
    summary.append(["Status Category", "Count"])

    for row in summary["A"]:
        row.font = Font(bold=True)

    for name in order:
        summary.append([name, len(grouped[name])])

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 70

    for name in order:
        ws = wb.create_sheet(name[:31])
        write_sheet(ws, grouped[name])

    wb.save(xlsx_file)

    headers = [
        "Page URL", "Resource URL", "Resource Type", "HTTP Status",
        "Status Category", "Final URL", "Content-Type",
        "Response Time (s)", "Error"
    ]
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    print("\nDone!")
    print(f"Excel: {xlsx_file}")
    print(f"CSV:   {csv_file}")
    print("\nSummary:")
    for name in order:
        print(f"  {name}: {len(grouped[name])}")


if __name__ == "__main__":
    main()
