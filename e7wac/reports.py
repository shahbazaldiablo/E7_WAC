import csv
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from .config import STATUS_ORDER

def write_excel_sheet(ws, rows, headers=None, col_widths=None):
    if headers is None:
        headers = ["Requested URL", "Resource URL", "Resource Type", "Initial Status", "Final Status", 
                   "Status Category", "Redirect Chain", "Final URL", "Content-Type", "Response Time (s)", "Error", "Classification", "Severity"]
    if col_widths is None:
        col_widths = [45, 65, 22, 13, 13, 22, 65, 65, 35, 20, 55, 30, 15]
        
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        
    if len(rows) > 0:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def generate_reports_from_db(db_name, root, stamp, safe_name, elapsed, folder_name, mode):
    conn = sqlite3.connect(db_name)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    xlsx_filename = os.path.join(folder_name, f"{safe_name}_website_audit_v0.8.0_{stamp}.xlsx")
    csv_filename = xlsx_filename.replace(".xlsx", ".csv")
    
    c.execute("SELECT COUNT(*) FROM discovered_pages")
    discovered_cnt = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM checked_pages")
    crawled_cnt = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM checked_resources")
    res_cnt = c.fetchone()[0]
    
    wb = Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = [
        ["Website", root],
        ["Scan timestamp", stamp],
        ["Audit Mode", mode.upper()],
        ["Runtime (minutes)", round(elapsed / 60, 2)],
        ["Pages discovered", discovered_cnt],
        ["Pages crawled", crawled_cnt],
        ["Resources checked", res_cnt],
        [], ["Status Category", "Count"]
    ]
    for row_data in summary_data: ws_summary.append(row_data)
        
    if mode != "images":
        for name in STATUS_ORDER:
            c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='Standard' AND status_category=?", (name,))
            row = c.fetchone()
            ws_summary.append([name, f"{row[0]} occurrences (Unique: {row[1]})"])
            
        c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='External'")
        row = c.fetchone()
        ws_summary.append(["External Links", f"{row[0]} occurrences (Unique: {row[1]})"])
        
        c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='WP Technical'")
        row = c.fetchone()
        ws_summary.append(["WP/Technical Resources", f"{row[0]} occurrences (Unique: {row[1]})"])
        
        c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='Sitemap'")
        row = c.fetchone()
        ws_summary.append(["Sitemaps", f"{row[0]} occurrences (Unique: {row[1]})"])
        
        c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='Potential legacy/staging domain'")
        row = c.fetchone()
        ws_summary.append(["Potential Legacy/Staging", f"{row[0]} occurrences (Unique: {row[1]})"])
        
        ws_summary.append([])
        ws_summary.append(["WordPress REST Post Type", "Public URLs discovered"])
        wp_counts = c.execute("SELECT typ, count FROM meta_wp ORDER BY typ").fetchall()
        for typ, count in wp_counts: ws_summary.append([typ, count])
            
        ws_summary.column_dimensions["A"].width = 32
        ws_summary.column_dimensions["B"].width = 70
        
        def get_rows(query, params=()):
            c.execute(query, params)
            return [{"Requested URL": r["page_url"], "Resource URL": r["resource_url"], "Resource Type": r["resource_type"],
                     "Initial Status": r["initial_status"], "Final Status": r["final_status"], "Status Category": r["status_category"],
                     "Redirect Chain": r["redirect_chain"], "Final URL": r["final_url"], "Content-Type": r["content_type"],
                     "Response Time (s)": r["response_time"], "Error": r["error"], "Classification": r["classification"], "Severity": r["severity"],
                     "Title": r["seo_title"], "Meta Description": r["seo_desc"], "Canonical": r["seo_canonical"],
                     "Robots": r["seo_robots"], "H1": r["seo_h1"]} for r in c.fetchall()]

        for name in STATUS_ORDER:
            rows = get_rows("SELECT * FROM results WHERE classification='Standard' AND status_category=?", (name,))
            if rows or name == "200 OK":
                write_excel_sheet(wb.create_sheet(name), rows)
                
        for cls_name, tab_name in [("External", "External Links"), ("WP Technical", "WP Technical"), ("Sitemap", "Sitemaps"), ("Potential legacy/staging domain", "Potential Staging Domains")]:
            rows = get_rows("SELECT * FROM results WHERE classification=?", (cls_name,))
            if rows: write_excel_sheet(wb.create_sheet(tab_name), rows)

        c.execute("""
            SELECT resource_url, initial_status, status_category, resource_type, severity, COUNT(*) as occurrence_count, MIN(page_url) as first_found_on 
            FROM results 
            WHERE initial_status >= 400 OR initial_status == 0 
            GROUP BY resource_url 
            ORDER BY occurrence_count DESC
        """)
        unique_rows = [{"Resource URL": r[0], "HTTP Status": r[1], "Status Category": r[2], "Resource Type": r[3], "Severity": r[4], "Occurrence Count": r[5], "First Found On": r[6]} for r in c.fetchall()]
        if unique_rows:
            write_excel_sheet(wb.create_sheet("Unique Issues"), unique_rows, 
                              headers=["Resource URL", "HTTP Status", "Status Category", "Resource Type", "Severity", "Occurrence Count", "First Found On"],
                              col_widths=[65, 13, 22, 22, 15, 18, 65])

        if mode in ["full", "seo"]:
            seo_headers = ["Page URL", "Initial Status", "Severity", "Title", "Meta Description", "Canonical", "Robots", "H1"]
            seo_widths = [55, 13, 15, 60, 80, 55, 30, 45]
            c.execute("SELECT page_url, initial_status, severity, seo_title, seo_desc, seo_canonical, seo_robots, seo_h1 FROM results WHERE resource_type='Page' AND initial_status=200")
            seo_rows = [{"Page URL": r[0], "Initial Status": r[1], "Severity": r[2], "Title": r[3], "Meta Description": r[4], "Canonical": r[5], "Robots": r[6], "H1": r[7]} for r in c.fetchall()]
            write_excel_sheet(wb.create_sheet("On-Page SEO"), seo_rows, seo_headers, seo_widths)

    else:
        # IMAGES MODE SPECIFIC TABS
        img_headers = ["Page URL", "Image URL", "HTTP Status", "Status Category", "Final URL", "Content-Type", "Error", "Redirect Chain"]
        img_widths = [55, 65, 13, 22, 65, 30, 45, 65]
        
        c.execute("SELECT page_url, resource_url, initial_status, status_category, final_url, content_type, error, redirect_chain FROM results WHERE resource_type != 'Page'")
        all_imgs = [{"Page URL": r[0], "Image URL": r[1], "HTTP Status": r[2], "Status Category": r[3], "Final URL": r[4], "Content-Type": r[5], "Error": r[6], "Redirect Chain": r[7]} for r in c.fetchall()]
        write_excel_sheet(wb.create_sheet("Image Audit"), all_imgs, img_headers, img_widths)
        
        broken_imgs = [r for r in all_imgs if r["HTTP Status"] >= 400 or r["HTTP Status"] == 0]
        write_excel_sheet(wb.create_sheet("Broken Images"), broken_imgs, img_headers, img_widths)
        
        c.execute("""
            SELECT resource_url, initial_status, status_category, COUNT(*) as occurrence_count, MIN(page_url) as first_found_on 
            FROM results 
            WHERE (initial_status >= 400 OR initial_status == 0) AND resource_type != 'Page'
            GROUP BY resource_url 
            ORDER BY occurrence_count DESC
        """)
        unique_img_rows = [{"Image URL": r[0], "HTTP Status": r[1], "Status Category": r[2], "Occurrence Count": r[3], "First Found On": r[4]} for r in c.fetchall()]
        write_excel_sheet(wb.create_sheet("Unique Image Issues"), unique_img_rows, 
                          headers=["Image URL", "HTTP Status", "Status Category", "Occurrence Count", "First Found On"],
                          col_widths=[65, 13, 22, 18, 65])
                          
        c.execute("SELECT page_url, resource_url, initial_status, status_category, classification FROM results WHERE classification='Potential legacy/staging domain' AND resource_type != 'Page'")
        staging_imgs = [{"Source page": r[0], "Image URL": r[1], "Status": r[2], "Detected reason/category": r[4]} for r in c.fetchall()]
        write_excel_sheet(wb.create_sheet("Potential Legacy Staging Images"), staging_imgs, 
                          headers=["Source page", "Image URL", "Status", "Detected reason/category"],
                          col_widths=[55, 65, 13, 30])

    wb.save(xlsx_filename)
    
    # CSV Writer
    csv_headers = ["Requested URL", "Resource URL", "Resource Type", "Initial Status", "Final Status", "Status Category", "Redirect Chain", "Final URL", "Content-Type", "Response Time (s)", "Error", "Classification", "Severity", "Title", "Meta Description", "Canonical", "Robots", "H1"]
    with open(csv_filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers)
        writer.writeheader()
        c.execute("SELECT * FROM results")
        for r in c.fetchall():
            writer.writerow({
                "Requested URL": r["page_url"], "Resource URL": r["resource_url"], "Resource Type": r["resource_type"],
                "Initial Status": r["initial_status"], "Final Status": r["final_status"], "Status Category": r["status_category"], 
                "Redirect Chain": r["redirect_chain"], "Final URL": r["final_url"],
                "Content-Type": r["content_type"], "Response Time (s)": r["response_time"], "Error": r["error"],
                "Classification": r["classification"], "Severity": r["severity"],
                "Title": r["seo_title"], "Meta Description": r["seo_desc"], "Canonical": r["seo_canonical"],
                "Robots": r["seo_robots"], "H1": r["seo_h1"]
            })
            
    print(f"\nExcel: {xlsx_filename}\nCSV:   {csv_filename}")
    conn.close()
