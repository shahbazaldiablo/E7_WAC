#!/usr/bin/env python3
import csv, re, sys, time
from collections import defaultdict, deque
from datetime import datetime
from urllib.parse import urljoin, urlparse, urldefrag

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEFAULT_URL = "https://indyroofandrestoration.com/"
UA = "Eagle7 Website Audit Crawler/2.0"
TIMEOUT, DELAY = 20, 0.10
MAX_PAGES, MAX_RESOURCES, API_PER_PAGE = 10000, 50000, 100
STATUS_ORDER = ["200 OK","3xx Redirects","4xx Client Errors","5xx Server Errors","Connection Errors","Other"]

def clean(u):
    if not u: return ""
    u, _ = urldefrag(u.strip())
    return u

def same(u, netloc):
    p = urlparse(u)
    return p.scheme in ("http","https") and p.netloc.lower() == netloc.lower()

def cat(s):
    if s == 200: return "200 OK"
    if s == 0: return "Connection Errors"
    if 300 <= s < 400: return "3xx Redirects"
    if 400 <= s < 500: return "4xx Client Errors"
    if 500 <= s < 600: return "5xx Server Errors"
    return "Other"

def suspicious(u, root_netloc):
    host = urlparse(u).netloc.lower()
    if host == root_netloc.lower(): return False
    text = host + urlparse(u).path.lower()
    return any(re.search(p, text) for p in [
        r"\bstaging\b",r"\bstage\b",r"\bdev\b",r"\bdevelopment\b",
        r"\btest\b",r"\btesting\b",r"\bqa\b",r"\buat\b",r"\bpreview\b",
        r"\blocalhost\b",r"\b127\.0\.0\.1\b",r"\.local\b",r"\.dev\b"
    ])

def css_urls(text):
    return [x.strip().strip("'\"") for x in re.findall(r"url\(\s*([^)]*)\)", text or "", re.I)
            if x.strip() and not x.strip().lower().startswith("data:")]

def discover(page, html):
    soup, found = BeautifulSoup(html, "html.parser"), []
    def add(raw, kind):
        if not raw or raw.strip().startswith(("data:","javascript:","mailto:","tel:","#")): return
        u = clean(urljoin(page, raw.strip()))
        if urlparse(u).scheme in ("http","https"): found.append((u,kind))
    for t in soup.find_all("a", href=True): add(t["href"], "Link")
    for t in soup.find_all("img"):
        add(t.get("src"), "Image")
        for x in (t.get("srcset") or "").split(","):
            if x.strip(): add(x.strip().split()[0], "Image srcset")
    for t in soup.find_all("source"):
        add(t.get("src"), "Source")
        for x in (t.get("srcset") or "").split(","):
            if x.strip(): add(x.strip().split()[0], "Source srcset")
    for t in soup.find_all("script", src=True): add(t["src"], "Script")
    for t in soup.find_all("link", href=True):
        add(t["href"], "Stylesheet" if "stylesheet" in " ".join(t.get("rel",[])).lower() else "Link resource")
    for t in soup.find_all(style=True):
        for x in css_urls(t.get("style")): add(x, "CSS URL")
    for x in css_urls(html): add(x, "CSS URL")
    return found

def req(s, u):
    started = time.perf_counter()
    try:
        r = s.get(u, timeout=TIMEOUT, allow_redirects=True, stream=True)
        out = {"status":r.status_code,"final_url":r.url,"content_type":r.headers.get("Content-Type",""),
               "response_time":round(time.perf_counter()-started,3),"error":""}
        r.close(); return out
    except requests.RequestException as e:
        return {"status":0,"final_url":"","content_type":"","response_time":round(time.perf_counter()-started,3),"error":str(e)}

def html(s,u):
    try:
        r=s.get(u,timeout=TIMEOUT,allow_redirects=True)
        out=(r.text if "text/html" in r.headers.get("Content-Type","").lower() else None,r.url,r.status_code,r.headers.get("Content-Type",""))
        r.close(); return out
    except requests.RequestException:
        return None,"",0,""

def discover_wp(s, root, netloc, add_page):
    base=urljoin(root,"/wp-json/wp/v2/")
    counts={}
    try:
        r=s.get(urljoin(base,"types"),timeout=TIMEOUT)
        if r.status_code != 200: r.close(); return counts
        types=r.json(); r.close()
    except (requests.RequestException,ValueError):
        return counts
    endpoints=[]
    for slug,info in types.items():
        if info.get("rest_base"): endpoints.append((slug,info["rest_base"]))
    known={x[1] for x in endpoints}
    if "posts" not in known: endpoints.append(("post","posts"))
    if "pages" not in known: endpoints.append(("page","pages"))
    print(f"  [WP API] Found {len(endpoints)} public REST post types")
    for typ,base_name in endpoints:
        endpoint=urljoin(base,base_name); page_no=1; count=0
        while True:
            try:
                r=s.get(endpoint,params={"per_page":API_PER_PAGE,"page":page_no,"_fields":"id,link,type,slug,status"},timeout=TIMEOUT)
                if r.status_code == 400 and page_no > 1: r.close(); break
                if r.status_code != 200: r.close(); break
                data=r.json(); total=int(r.headers.get("X-WP-TotalPages","1")); r.close()
            except (requests.RequestException,ValueError): break
            for item in data:
                u=clean(item.get("link",""))
                if u and same(u,netloc): add_page(u, f"WP REST: {typ}"); count+=1
            if page_no >= total: break
            page_no += 1
        counts[typ]=count
        print(f"  [WP API] {typ}: {count} public URLs")
    return counts

def write_sheet(ws, rows):
    headers=["Page URL","Resource URL","Resource Type","HTTP Status","Status Category","Final URL","Content-Type","Response Time (s)","Error"]
    ws.append(headers)
    for c in ws[1]:
        c.font=Font(bold=True); c.fill=PatternFill("solid",fgColor="D9EAF7")
    for row in rows: ws.append([row.get(h,"") for h in headers])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for i,w in enumerate([45,65,22,13,22,65,35,20,55],1): ws.column_dimensions[chr(64+i)].width=w

def main():
    root=sys.argv[1] if len(sys.argv)>1 else DEFAULT_URL
    if not root.startswith(("http://","https://")): root="https://"+root
    root=clean(root).rstrip("/")+"/"; netloc=urlparse(root).netloc
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); safe=re.sub(r"[^A-Za-z0-9.-]","_",netloc)
    xlsx=f"{safe}_website_audit_v2_{stamp}.xlsx"; csvfile=xlsx.replace(".xlsx",".csv")
    s=requests.Session(); s.headers.update({"User-Agent":UA,"Accept":"*/*","Accept-Language":"en-GB,en;q=0.9"})
    queue=deque([root]); discovered={root}; checked_pages=set(); checked_resources=set(); rows=[]; tech=[]; suspicious_rows=[]
    started=time.perf_counter()
    print("\nEagle7 Website Audit Crawler v2\n"+"="*60+f"\nTarget: {root}\n"+"="*60)
    def add_page(u, source="Front-end"):
        u=clean(u)
        if u and same(u,netloc) and u not in discovered:
            discovered.add(u); queue.append(u)
    print("\n[PHASE 1] Discovering WordPress content...")
    cpt=discover_wp(s,root,netloc,add_page)
    print(f"\nTotal pages discovered so far: {len(discovered)}")
    print("\n[PHASE 2] Crawling pages and checking resources...\n")
    while queue and len(checked_pages)<MAX_PAGES:
        page=queue.popleft()
        if page in checked_pages: continue
        checked_pages.add(page)
        elapsed=time.perf_counter()-started
        rate=len(checked_pages)/elapsed*60 if elapsed else 0
        denominator=max(len(discovered),len(checked_pages))
        pct=len(checked_pages)/denominator*100 if denominator else 0
        eta=(denominator-len(checked_pages))/rate if rate else 0
        print(f"[{pct:6.2f}%] Pages {len(checked_pages)}/{denominator} | Discovered {len(discovered)} | Resources {len(checked_resources)} | {rate:.1f} pages/min | ETA {eta:.1f} min")
        print(f"         {page}")
        h,final,_,_=html(s,page); result=req(s,page)
        row={"Page URL":page,"Resource URL":page,"Resource Type":"Page","HTTP Status":result["status"],"Status Category":cat(result["status"]),"Final URL":result["final_url"],"Content-Type":result["content_type"],"Response Time (s)":result["response_time"],"Error":result["error"]}
        rows.append(row)
        if suspicious(page,netloc): suspicious_rows.append({**row,"Issue":"Potential staging/development/local URL"})
        if h is None: time.sleep(DELAY); continue
        for u,kind in discover(final or page,h):
            path=urlparse(u).path.lower()
            if kind=="Link" and same(u,netloc) and not path.startswith("/wp-json/") and "/oembed/" not in path:
                if u not in discovered: discovered.add(u); queue.append(u)
            key=(u,kind)
            if key in checked_resources or len(checked_resources)>=MAX_RESOURCES: continue
            checked_resources.add(key); result=req(s,u)
            row={"Page URL":page,"Resource URL":u,"Resource Type":kind,"HTTP Status":result["status"],"Status Category":cat(result["status"]),"Final URL":result["final_url"],"Content-Type":result["content_type"],"Response Time (s)":result["response_time"],"Error":result["error"]}
            rows.append(row)
            if "/wp-json/" in u.lower() or "/oembed/" in u.lower(): tech.append(row)
            if suspicious(u,netloc): suspicious_rows.append({**row,"Issue":"Potential staging/development/local URL"})
            time.sleep(DELAY)
        time.sleep(DELAY)
    grouped=defaultdict(list)
    for r in rows: grouped[r["Status Category"]].append(r)
    for r in tech:
        if r in grouped[r["Status Category"]]: grouped[r["Status Category"]].remove(r)
    elapsed=time.perf_counter()-started
    print("\n"+"="*60+"\nSCAN COMPLETE\n"+"="*60)
    print(f"Pages crawled: {len(checked_pages)}\nPages discovered: {len(discovered)}\nResources checked: {len(checked_resources)}\nRuntime: {elapsed/60:.1f} minutes")
    for name in STATUS_ORDER: print(f"{name:22}: {len(grouped[name])}")
    print(f"{'WP/technical resources':22}: {len(tech)}\n{'Suspicious URLs':22}: {len(suspicious_rows)}")
    wb=Workbook(); wb.remove(wb.active)
    ws=wb.create_sheet("Summary")
    for row in [["Website",root],["Scan timestamp",stamp],["Runtime (minutes)",round(elapsed/60,2)],["Pages discovered",len(discovered)],["Pages crawled",len(checked_pages)],["Resources checked",len(checked_resources)],[],["Status Category","Count"]]:
        ws.append(row)
    for name in STATUS_ORDER: ws.append([name,len(grouped[name])])
    ws.append(["WP/Technical Resources",len(tech)]); ws.append(["Suspicious URLs",len(suspicious_rows)])
    ws.append([]); ws.append(["WordPress REST Post Type","Public URLs discovered"])
    for typ,count in sorted(cpt.items()): ws.append([typ,count])
    ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=70
    for name in STATUS_ORDER: write_sheet(wb.create_sheet(name),grouped[name])
    if tech: write_sheet(wb.create_sheet("WP Technical"),tech)
    if suspicious_rows:
        write_sheet(wb.create_sheet("Suspicious URLs"),suspicious_rows)
    wb.save(xlsx)
    headers=["Page URL","Resource URL","Resource Type","HTTP Status","Status Category","Final URL","Content-Type","Response Time (s)","Error"]
    with open(csvfile,"w",newline="",encoding="utf-8-sig") as f:
        w=csv.DictWriter(f,fieldnames=headers); w.writeheader(); w.writerows(rows)
    print(f"\nExcel: {xlsx}\nCSV:   {csvfile}")

if __name__=="__main__": main()
