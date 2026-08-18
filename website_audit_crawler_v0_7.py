#!/usr/bin/env python3
import csv
import os
import re
import sys
import time
import sqlite3
import msvcrt
from collections import deque
from datetime import datetime
from urllib.parse import urljoin, urlparse, urldefrag
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_URL = "https://eagle7.in"
USER_AGENT = "Eagle7 Website Audit Crawler/0.7"
TIMEOUT = 20
DELAY = 0.05
MAX_PAGES = 10000
MAX_RESOURCES = 50000
API_PER_PAGE = 100
MAX_WORKERS = 10
STATUS_ORDER = ["200 OK", "3xx Redirects", "4xx Client Errors", "5xx Server Errors", "Connection Errors", "Other"]
HEAVY_EXTENSIONS = {'.pdf', '.mp4', '.zip', '.gz', '.tar', '.rar', '.exe', '.dmg', '.iso', '.bin', '.apk', '.jpg', '.jpeg', '.png', '.gif', '.svg'}

is_paused = False


def check_controls():
    global is_paused
    while msvcrt.kbhit():
        char = msvcrt.getch()
        if char.lower() == b'p':
            is_paused = not is_paused
            if is_paused:
                print("\n" + "="*60 + "\nSCAN PAUSED\n" + "="*60)
                print("Press P to resume\nPress Ctrl+C to stop and save\n" + "="*60)
            else:
                print("\n" + "="*60 + "\nSCAN RESUMED\n" + "="*60)


def clean_url(u):
    if not u: return ""
    u, _ = urldefrag(u.strip())
    parsed = urlparse(u)
    if parsed.scheme:
        u = parsed._replace(scheme=parsed.scheme.lower(), netloc=parsed.netloc.lower()).geturl()
    if parsed.path == "" and parsed.netloc:
        u += "/"
    return u


def is_same_domain(u, root_netloc):
    parsed = urlparse(u)
    return parsed.scheme in ("http", "https") and parsed.netloc.lower() == root_netloc.lower()


def categorize_status(status_code):
    if status_code == 200: return "200 OK"
    if status_code == 0: return "Connection Errors"
    if 300 <= status_code < 400: return "3xx Redirects"
    if 400 <= status_code < 500: return "4xx Client Errors"
    if 500 <= status_code < 600: return "5xx Server Errors"
    return "Other"


def get_classification(url, root_netloc, kind=""):
    path = urlparse(url).path.lower()
    host = urlparse(url).netloc.lower()
    
    if path.endswith(".xml") or "sitemap" in path:
        return "Sitemap"
        
    if "/wp-json/" in path or "/oembed/" in path:
        return "WP Technical"
        
    if host != root_netloc.lower():
        text = host + path
        patterns = [
            r"\bstaging\b", r"\bstage\b", r"\bdev\b", r"\bdevelopment\b",
            r"\btest\b", r"\btesting\b", r"\bqa\b", r"\buat\b", r"\bpreview\b",
            r"\blocalhost\b", r"\b127\.0\.0\.1\b", r"\.local\b", r"\.dev\b"
        ]
        if any(re.search(p, text) for p in patterns):
            return "Potential legacy/staging domain"
        return "External"
        
    return "Standard"


def extract_css_urls(text):
    if not text: return []
    return [match.strip().strip("'\"") for match in re.findall(r"url\(\s*([^)]*)\)", text, re.IGNORECASE)
            if match.strip() and not match.strip().lower().startswith("data:")]


def extract_seo_tags(html_content):
    if not html_content:
        return {"seo_title": "", "seo_desc": "", "seo_canonical": "", "seo_robots": "", "seo_h1": ""}
    
    soup = BeautifulSoup(html_content, "html.parser")
    title = soup.find("title")
    desc_tag = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
    canonical_tag = soup.find("link", rel=re.compile(r"^canonical$", re.I))
    robots_tag = soup.find("meta", attrs={"name": re.compile(r"^robots$", re.I)})
    h1_tag = soup.find("h1")
    
    h1_text = h1_tag.text.strip() if h1_tag else ""
    all_h1s = soup.find_all("h1")
    if len(all_h1s) > 1:
        h1_text = f"[MULTIPLE ({len(all_h1s)})] " + h1_text
        
    return {
        "seo_title": title.text.strip() if title else "",
        "seo_desc": desc_tag.get("content", "").strip() if desc_tag else "",
        "seo_canonical": canonical_tag.get("href", "").strip() if canonical_tag else "",
        "seo_robots": robots_tag.get("content", "").strip() if robots_tag else "",
        "seo_h1": h1_text
    }


def discover_assets(page_url, html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    found = []

    def add_asset(raw_url, kind):
        if not raw_url: return
        raw = raw_url.strip()
        if raw.lower().startswith(("data:", "javascript:", "mailto:", "tel:", "#")): return
        u = clean_url(urljoin(page_url, raw))
        if urlparse(u).scheme in ("http", "https"):
            found.append((u, kind))

    for tag in soup.find_all("a", href=True): add_asset(tag["href"], "Link")
    for tag in soup.find_all("img"):
        add_asset(tag.get("src"), "Image")
        for x in (tag.get("srcset") or "").split(","):
            if x.strip(): add_asset(x.strip().split()[0], "Image srcset")
    for tag in soup.find_all("source"):
        add_asset(tag.get("src"), "Source")
        for x in (tag.get("srcset") or "").split(","):
            if x.strip(): add_asset(x.strip().split()[0], "Source srcset")
    for tag in soup.find_all("script", src=True): add_asset(tag["src"], "Script")
    for tag in soup.find_all("link", href=True):
        rel = " ".join(tag.get("rel", [])).lower()
        add_asset(tag["href"], "Stylesheet" if "stylesheet" in rel else "Link resource")
    for tag in soup.find_all(style=True):
        for x in extract_css_urls(tag.get("style")): add_asset(x, "CSS URL")
    for tag in soup.find_all("style"):
        for x in extract_css_urls(tag.string): add_asset(x, "CSS URL")

    return found


def fetch_url(session, url, root_netloc, stream=False, is_page=False):
    started = time.perf_counter()
    current_url = url
    redirect_chain = []
    initial_status = None
    final_status = 0
    final_url = url
    content_type = ""
    html_text = None
    error_msg = ""
    
    try:
        while True:
            path = urlparse(current_url).path.lower()
            ext = path[path.rfind('.'):] if '.' in path else ''
            is_heavy = ext in HEAVY_EXTENSIONS
            is_external = not is_same_domain(current_url, root_netloc)
            
            req_kwargs = {"timeout": TIMEOUT, "allow_redirects": False}
            if not is_page and (is_heavy or is_external):
                r = session.head(current_url, **req_kwargs)
                if r.status_code == 405:
                    r = session.get(current_url, stream=True, **req_kwargs)
            else:
                r = session.get(current_url, stream=stream, **req_kwargs)
                
            if initial_status is None:
                initial_status = r.status_code
                
            if 300 <= r.status_code < 400 and 'Location' in r.headers:
                next_url = urljoin(current_url, r.headers['Location'])
                redirect_chain.append(current_url)
                current_url = clean_url(next_url)
                if len(redirect_chain) > 10:
                    error_msg = "Redirect Loop Detected"
                    break
            else:
                final_status = r.status_code
                final_url = r.url if hasattr(r, 'url') else current_url
                content_type = r.headers.get("Content-Type", "")
                if not stream and "text/html" in content_type.lower() and hasattr(r, 'text'):
                    html_text = r.text
                break
                
        redirect_chain_str = " -> ".join(redirect_chain + [final_url]) if redirect_chain else ""
        return {
            "initial_status": initial_status or 0, "final_status": final_status, 
            "final_url": final_url, "redirect_chain": redirect_chain_str, 
            "content_type": content_type, "response_time": round(time.perf_counter() - started, 3), 
            "error": error_msg, "html": html_text
        }
    except requests.RequestException as e:
        return {
            "initial_status": 0, "final_status": 0, "final_url": "", "redirect_chain": "",
            "content_type": "", "response_time": round(time.perf_counter() - started, 3), 
            "error": str(e), "html": None
        }


def discover_wp_and_sitemaps(session, root_url, netloc, add_page_callback, c):
    robots_url = urljoin(root_url, "/robots.txt")
    sitemaps = []
    try:
        r = session.get(robots_url, timeout=TIMEOUT)
        if r.status_code == 200:
            for line in r.text.splitlines():
                if line.lower().startswith("sitemap:"):
                    sitemaps.append(line.split(":", 1)[1].strip())
    except Exception:
        pass
        
    if not sitemaps:
        sitemaps.extend([urljoin(root_url, "/sitemap.xml"), urljoin(root_url, "/sitemap_index.xml")])
        
    print(f"  [SEO] Checking {len(sitemaps)} potential sitemaps...")
    for sm in sitemaps:
        try:
            r = session.get(sm, timeout=TIMEOUT)
            if r.status_code == 200:
                for loc in re.findall(r"<loc>(.*?)</loc>", r.text, re.IGNORECASE):
                    u = clean_url(loc)
                    if is_same_domain(u, netloc):
                        add_page_callback(u, "queue")
        except Exception:
            pass

    base_api = urljoin(root_url, "/wp-json/wp/v2/")
    try:
        r = session.get(urljoin(base_api, "types"), timeout=TIMEOUT)
        if r.status_code != 200: return
        types_data = r.json()
    except Exception:
        return

    endpoints = [(slug, info["rest_base"]) for slug, info in types_data.items() if info.get("rest_base")]
    known_bases = {x[1] for x in endpoints}
    if "posts" not in known_bases: endpoints.append(("post", "posts"))
    if "pages" not in known_bases: endpoints.append(("page", "pages"))
    
    print(f"  [WP API] Found {len(endpoints)} public REST post types")
    for typ, base_name in endpoints:
        endpoint = urljoin(base_api, base_name)
        page_no, count = 1, 0
        while True:
            check_controls()
            if is_paused:
                time.sleep(0.5)
                continue
                
            try:
                params = {"per_page": API_PER_PAGE, "page": page_no, "_fields": "id,link,type,slug,status"}
                r = session.get(endpoint, params=params, timeout=TIMEOUT)
                if r.status_code != 200 or (r.status_code == 400 and page_no > 1): break
                data = r.json()
                total_pages = int(r.headers.get("X-WP-TotalPages", "1"))
            except Exception: break
                
            for item in data:
                u = clean_url(item.get("link", ""))
                if u and is_same_domain(u, netloc):
                    add_page_callback(u, "queue")
                    count += 1
            if page_no >= total_pages: break
            page_no += 1
            
        print(f"  [WP API] {typ}: {count} public URLs")
        c.execute("INSERT OR REPLACE INTO meta_wp VALUES (?, ?)", (typ, count))


def write_excel_sheet(ws, rows, headers=None, col_widths=None):
    if headers is None:
        headers = ["Requested URL", "Resource URL", "Resource Type", "Initial Status", "Final Status", 
                   "Status Category", "Redirect Chain", "Final URL", "Content-Type", "Response Time (s)", "Error"]
    if col_widths is None:
        col_widths = [45, 65, 22, 13, 13, 22, 65, 65, 35, 20, 55]
        
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        
    if len(rows) > 0:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_reports_from_db(db_name, root, stamp, safe_name, elapsed, folder_name):
    conn = sqlite3.connect(db_name)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    xlsx_filename = os.path.join(folder_name, f"{safe_name}_website_audit_v0_7_{stamp}.xlsx")
    csv_filename = xlsx_filename.replace(".xlsx", ".csv")
    
    c.execute("SELECT COUNT(*) FROM discovered_pages")
    discovered_cnt = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM checked_pages")
    crawled_cnt = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM checked_resources")
    res_cnt = c.fetchone()[0]
    
    wb = Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = [
        ["Website", root],
        ["Scan timestamp", stamp],
        ["Runtime (minutes)", round(elapsed / 60, 2)],
        ["Pages discovered", discovered_cnt],
        ["Pages crawled", crawled_cnt],
        ["Resources checked", res_cnt],
        [], ["Status Category", "Count"]
    ]
    for row_data in summary_data: ws_summary.append(row_data)
        
    for name in STATUS_ORDER:
        c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='Standard' AND status_category=?", (name,))
        row = c.fetchone()
        ws_summary.append([name, f"{row[0]} occurrences (Unique: {row[1]})"])
        
    c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='External'")
    row = c.fetchone()
    ws_summary.append(["External Links", f"{row[0]} occurrences (Unique: {row[1]})"])
    
    c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='WP Technical'")
    row = c.fetchone()
    ws_summary.append(["WP/Technical Resources", f"{row[0]} occurrences (Unique: {row[1]})"])
    
    c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='Sitemap'")
    row = c.fetchone()
    ws_summary.append(["Sitemaps", f"{row[0]} occurrences (Unique: {row[1]})"])
    
    c.execute("SELECT COUNT(*), COUNT(DISTINCT resource_url) FROM results WHERE classification='Potential legacy/staging domain'")
    row = c.fetchone()
    ws_summary.append(["Potential Legacy/Staging", f"{row[0]} occurrences (Unique: {row[1]})"])
    
    ws_summary.append([])
    ws_summary.append(["WordPress REST Post Type", "Public URLs discovered"])
    wp_counts = c.execute("SELECT typ, count FROM meta_wp ORDER BY typ").fetchall()
    for typ, count in wp_counts: ws_summary.append([typ, count])
        
    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 70
    
    # Standard query template
    def get_rows(query, params=()):
        c.execute(query, params)
        return [{"Requested URL": r["page_url"], "Resource URL": r["resource_url"], "Resource Type": r["resource_type"],
                 "Initial Status": r["initial_status"], "Final Status": r["final_status"], "Status Category": r["status_category"],
                 "Redirect Chain": r["redirect_chain"], "Final URL": r["final_url"], "Content-Type": r["content_type"],
                 "Response Time (s)": r["response_time"], "Error": r["error"],
                 "Title": r["seo_title"], "Meta Description": r["seo_desc"], "Canonical": r["seo_canonical"],
                 "Robots": r["seo_robots"], "H1": r["seo_h1"]} for r in c.fetchall()]

    for name in STATUS_ORDER:
        rows = get_rows("SELECT * FROM results WHERE classification='Standard' AND status_category=?", (name,))
        if rows or name == "200 OK":
            write_excel_sheet(wb.create_sheet(name), rows)
            
    # Classifications
    for cls_name, tab_name in [("External", "External Links"), ("WP Technical", "WP Technical"), ("Sitemap", "Sitemaps"), ("Potential legacy/staging domain", "Potential Staging Domains")]:
        rows = get_rows("SELECT * FROM results WHERE classification=?", (cls_name,))
        if rows: write_excel_sheet(wb.create_sheet(tab_name), rows)

    # Unique Issues Tab
    c.execute("""
        SELECT resource_url, initial_status, status_category, resource_type, COUNT(*) as occurrence_count, MIN(page_url) as first_found_on 
        FROM results 
        WHERE initial_status >= 400 OR initial_status == 0 
        GROUP BY resource_url 
        ORDER BY occurrence_count DESC
    """)
    unique_rows = [{"Resource URL": r[0], "HTTP Status": r[1], "Status Category": r[2], "Resource Type": r[3], "Occurrence Count": r[4], "First Found On": r[5]} for r in c.fetchall()]
    if unique_rows:
        write_excel_sheet(wb.create_sheet("Unique Issues"), unique_rows, 
                          headers=["Resource URL", "HTTP Status", "Status Category", "Resource Type", "Occurrence Count", "First Found On"],
                          col_widths=[65, 13, 22, 22, 18, 65])

    # On-Page SEO Tab
    seo_headers = ["Page URL", "Initial Status", "Title", "Meta Description", "Canonical", "Robots", "H1"]
    seo_widths = [55, 13, 60, 80, 55, 30, 45]
    c.execute("SELECT page_url, initial_status, seo_title, seo_desc, seo_canonical, seo_robots, seo_h1 FROM results WHERE resource_type='Page' AND initial_status=200")
    seo_rows = [{"Page URL": r[0], "Initial Status": r[1], "Title": r[2], "Meta Description": r[3], "Canonical": r[4], "Robots": r[5], "H1": r[6]} for r in c.fetchall()]
    write_excel_sheet(wb.create_sheet("On-Page SEO"), seo_rows, seo_headers, seo_widths)

    wb.save(xlsx_filename)
    
    # CSV Writer
    csv_headers = ["Requested URL", "Resource URL", "Resource Type", "Initial Status", "Final Status", "Status Category", "Redirect Chain", "Final URL", "Content-Type", "Response Time (s)", "Error", "Classification", "Title", "Meta Description", "Canonical", "Robots", "H1"]
    with open(csv_filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers)
        writer.writeheader()
        c.execute("SELECT * FROM results")
        for r in c.fetchall():
            writer.writerow({
                "Requested URL": r["page_url"], "Resource URL": r["resource_url"], "Resource Type": r["resource_type"],
                "Initial Status": r["initial_status"], "Final Status": r["final_status"], "Status Category": r["status_category"], 
                "Redirect Chain": r["redirect_chain"], "Final URL": r["final_url"],
                "Content-Type": r["content_type"], "Response Time (s)": r["response_time"], "Error": r["error"],
                "Classification": r["classification"],
                "Title": r["seo_title"], "Meta Description": r["seo_desc"], "Canonical": r["seo_canonical"],
                "Robots": r["seo_robots"], "H1": r["seo_h1"]
            })
            
    print(f"\nExcel: {xlsx_filename}\nCSV:   {csv_filename}")
    conn.close()


def check_resource_task(session, url, kind, page_url, netloc):
    result = fetch_url(session, url, netloc, stream=True, is_page=False)
    classification = get_classification(url, netloc, kind)
    return {
        "page_url": page_url, "resource_url": url, "resource_type": kind,
        "initial_status": result["initial_status"], "final_status": result["final_status"],
        "status_category": categorize_status(result["initial_status"]),
        "redirect_chain": result["redirect_chain"], "final_url": result["final_url"],
        "content_type": result["content_type"], "response_time": result["response_time"], 
        "error": result["error"], "classification": classification,
        "seo_title": "", "seo_desc": "", "seo_canonical": "", "seo_robots": "", "seo_h1": ""
    }


def main():
    root = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_URL
    if not root.startswith(("http://", "https://")): root = "https://" + root
    root = clean_url(root)
    if not root.endswith("/"): root += "/"
        
    netloc = urlparse(root).netloc
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9.-]", "_", netloc)
    
    folders = [e.name for e in os.scandir('.') if e.is_dir() and e.name.startswith(f"{safe_name}_website_audit_")]
    is_resume = False
    
    if folders:
        folder_name = sorted(folders)[-1]
        print(f"\n[!] Existing scan found.\nResuming: {folder_name}")
        is_resume = True
    else:
        folder_name = f"{safe_name}_website_audit_{stamp}"
        os.makedirs(folder_name, exist_ok=True)
        
    db_name = os.path.join(folder_name, f"{safe_name}_audit.db")
    conn = sqlite3.connect(db_name)
    c = conn.cursor()
    
    c.execute("CREATE TABLE IF NOT EXISTS queue (id INTEGER PRIMARY KEY AUTOINCREMENT, url TEXT UNIQUE)")
    c.execute("CREATE TABLE IF NOT EXISTS discovered_pages (url TEXT PRIMARY KEY)")
    c.execute("CREATE TABLE IF NOT EXISTS checked_pages (url TEXT PRIMARY KEY)")
    c.execute("CREATE TABLE IF NOT EXISTS checked_resources (url TEXT, kind TEXT, PRIMARY KEY(url, kind))")
    c.execute("""CREATE TABLE IF NOT EXISTS results (
        page_url TEXT, resource_url TEXT, resource_type TEXT, initial_status INTEGER, final_status INTEGER,
        status_category TEXT, redirect_chain TEXT, final_url TEXT, content_type TEXT, response_time REAL, 
        error TEXT, classification TEXT,
        seo_title TEXT, seo_desc TEXT, seo_canonical TEXT, seo_robots TEXT, seo_h1 TEXT
    )""")
    c.execute("CREATE TABLE IF NOT EXISTS meta_wp (typ TEXT PRIMARY KEY, count INTEGER)")
    conn.commit()
    
    discovered_pages = {r[0] for r in c.execute("SELECT url FROM discovered_pages").fetchall()}
    checked_pages = {r[0] for r in c.execute("SELECT url FROM checked_pages").fetchall()}
    checked_resources = {(r[0], r[1]) for r in c.execute("SELECT url, kind FROM checked_resources").fetchall()}
    
    queue_rows = c.execute("SELECT url FROM queue ORDER BY id ASC").fetchall()
    queue = deque([r[0] for r in queue_rows])
    
    def db_add_page(u, table="discovered_pages"):
        if table == "discovered_pages":
            c.execute("INSERT OR IGNORE INTO discovered_pages VALUES (?)", (u,))
        elif table == "queue":
            if u not in discovered_pages:
                discovered_pages.add(u)
                queue.append(u)
                c.execute("INSERT OR IGNORE INTO discovered_pages VALUES (?)", (u,))
                c.execute("INSERT OR IGNORE INTO queue (url) VALUES (?)", (u,))
    
    if not is_resume or not discovered_pages:
        db_add_page(root, "queue")
        conn.commit()
        
    retry_strategy = Retry(total=3, status_forcelist=[429, 500, 502, 503, 504], backoff_factor=1)
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*", "Accept-Language": "en-GB,en;q=0.9"})
    
    started = time.perf_counter()
    recent_completions = deque(maxlen=50)
    
    print("\nEagle7 Website Audit Crawler v0.7")
    print("=" * 60)
    print(f"Target: {root}")
    print("\nControls:")
    print("  P       -> Pause / Resume")
    print("  Ctrl+C  -> Stop and save safely")
    print("=" * 60)
    
    try:
        if not is_resume:
            print("\n[PHASE 1] Discovering content via WP REST, Sitemaps & Robots...")
            discover_wp_and_sitemaps(session, root, netloc, db_add_page, c)
            conn.commit()
            
        print(f"\n[PHASE 2] Crawling {len(discovered_pages)} discovered pages...\n")
        
        while queue and len(checked_pages) < MAX_PAGES:
            check_controls()
            if is_paused:
                time.sleep(0.5)
                continue
                
            page_url = queue.popleft()
            c.execute("DELETE FROM queue WHERE url=?", (page_url,))
            
            if page_url in checked_pages:
                conn.commit()
                continue
                
            checked_pages.add(page_url)
            c.execute("INSERT OR IGNORE INTO checked_pages VALUES (?)", (page_url,))
            
            recent_completions.append(time.perf_counter())
            if len(recent_completions) > 1:
                t_diff = recent_completions[-1] - recent_completions[0]
                rate_per_sec = (len(recent_completions) - 1) / t_diff if t_diff > 0 else 0
            else:
                rate_per_sec = 0
                
            rate_per_min = rate_per_sec * 60
            denominator = max(len(discovered_pages), len(checked_pages))
            eta = (denominator - len(checked_pages)) / rate_per_min if rate_per_min else 0
            
            if len(checked_pages) < 5 or rate_per_min == 0:
                eta_str = "Calculating..."
            else:
                eta_str = f"~{eta:.1f} m"
            
            print(f"Pages: {len(checked_pages)} crawled / {denominator} discovered | Resources: {len(checked_resources)} | {rate_per_min:.1f} p/min | ETA: {eta_str}")
            print(f"  -> {page_url}")
            
            result = fetch_url(session, page_url, netloc, stream=False, is_page=True)
            html_content = result["html"]
            seo_data = extract_seo_tags(html_content)
            
            row_tpl = (
                page_url, page_url, "Page", result["initial_status"], result["final_status"], categorize_status(result["initial_status"]),
                result["redirect_chain"], result["final_url"], result["content_type"], result["response_time"], result["error"],
                get_classification(page_url, netloc, "Page"),
                seo_data["seo_title"], seo_data["seo_desc"], seo_data["seo_canonical"], seo_data["seo_robots"], seo_data["seo_h1"]
            )
            c.execute("INSERT INTO results VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", row_tpl)
            
            if not html_content:
                conn.commit()
                time.sleep(DELAY)
                continue
                
            assets = discover_assets(result["final_url"] or page_url, html_content)
            tasks = []
            
            for u, kind in assets:
                path = urlparse(u).path.lower()
                if kind == "Link" and is_same_domain(u, netloc) and not path.startswith("/wp-json/") and "/oembed/" not in path and not (path.endswith(".xml") or "sitemap" in path):
                    db_add_page(u, "queue")
                
                key = (u, kind)
                if key in checked_resources or len(checked_resources) >= MAX_RESOURCES:
                    continue
                    
                checked_resources.add(key)
                tasks.append((u, kind))

            if tasks:
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    future_to_task = {executor.submit(check_resource_task, session, u, kind, page_url, netloc): u for u, kind in tasks}
                    for future in as_completed(future_to_task):
                        check_controls()
                        while is_paused:
                            check_controls()
                            time.sleep(0.5)
                        try:
                            r = future.result()
                            c.execute("INSERT OR IGNORE INTO checked_resources VALUES (?, ?)", (r["resource_url"], r["resource_type"]))
                            res_tpl = (
                                r["page_url"], r["resource_url"], r["resource_type"], r["initial_status"], r["final_status"], r["status_category"],
                                r["redirect_chain"], r["final_url"], r["content_type"], r["response_time"], r["error"],
                                r["classification"],
                                r["seo_title"], r["seo_desc"], r["seo_canonical"], r["seo_robots"], r["seo_h1"]
                            )
                            c.execute("INSERT INTO results VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", res_tpl)
                        except Exception:
                            pass
                            
            conn.commit()
            time.sleep(DELAY)
            
    except KeyboardInterrupt:
        print("\n" + "="*60)
        print("SCAN STOPPED SAFELY")
        print("="*60)
        print("\nThe scan was stopped by the user.")
        print("All completed results have been saved.")
        print("The scan can be resumed using the same database.\n")
        
    elapsed = time.perf_counter() - started
    if not is_paused and queue:
        # Partial generation
        print("Generating partial reports...")
    else:
        print("\n" + "=" * 60 + "\nSCAN COMPLETE\n" + "=" * 60)
        print("Generating Reports...")
        
    generate_reports_from_db(db_name, root, stamp, safe_name, elapsed, folder_name)

if __name__ == "__main__":
    main()
